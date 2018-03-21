#!/usr/bin/env python
#coding=utf-8

"""
Generates a binary strings for paradigms and writes it to a nexus file.

"""

import os
import binascii

try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError("Please install openpyxl")

try:
    from nexus import NexusWriter
except ImportError:
    raise ImportError("Please install python-nexus")

#open the xl file and read the rows into the sibling term dictionary

def get_data(xlsfile):
    wb = load_workbook(filename=xlsfile)
    w = wb.worksheets[0]
    siblings = []
    header = None
    sibline = {}
    for row in w.iter_rows():
        row = [r.value for r in row]
        if header is None:
            header = row
        else:
            sibline = dict(zip(header, row))
            sibline.pop(None,None)
        siblings.append(sibline)
            
    return siblings

# create a dictionary containing canonical representations of the paradigm cells
#
# vector is created using the following formula
# sex of speaker m=01 f=10
# ref generation 2=10000 1=01000 0=00100 -1=00010 -2=00001
# linking rel generation 2=10000 1=01000 0=00100 -1=00010 -2=00001
# rel age of link e=01 y=10 either=11 not relevant=00
# sex of ref m=01 f=10
# sex of link m=01 f=10
#  


canon = { 'fBW' : 0b10001000100000000000000000100100001,
          'fD' : 0b10000000000000000000000000010100000,
          'fDD' : 0b10000101000000000000000000001100000,
          'fDH' : 0b10000101000000000000000000010010001,
          'fDS' : 0b10000101000000000000000000001010000,
          'feB' : 0b10000000000000000000000000100011000,
          'feBD' : 0b10001000110000000000000000010100000,
          'feBS' : 0b10001000110000000000000000010010000,
          'feZ' : 0b10000000000000000000000000100101000,
          'feZD' : 0b10001001010000000000000000010100000,
          'feZS' : 0b10001001010000000000000000010010000,
          'fF' : 0b10000000000000000000000001000010000,
          'fFBD' : 0b10010000100000100001000000100100000,
          'fFBS' : 0b10010000100000100001000000100010000,
          'fFeB' : 0b10010000100000000000000001000011000,
          'fFeZ' : 0b10010000100000000000000001000101000,
          'fFF' : 0b10010000100000000000000010000010000,
          'fFM' : 0b10010000100000000000000010000100000,
          'fFyB' : 0b10010000100000000000000001000010100,
          'fFyZ' : 0b10010000100000000000000001000100100,
          'fFZD' : 0b10010000100000100010000000100100000,
          'fFZS' : 0b10010000100000100010000000100010000,
          'fH' : 0b10000000000000000000000000100010001,
          'fHB' : 0b10001000100010000000000000100010000,
          'fHF' : 0b10001000100010000000000001000010000,
          'fHM' : 0b10001000100010000000000001000100000,
          'fHZ' : 0b10001000100010000000000000100100000,
          'fM' : 0b10000000000000000000000001000100000,
          'fMBD' : 0b10010001000000100001000000100100000,
          'fMBS' : 0b10010001000000100001000000100010000,
          'fMeB' : 0b10010001000000000000000001000011000,
          'fMeZ' : 0b10010001000000000000000001000101000,
          'fMF' : 0b10010001000000000000000010000010000,
          'fMM' : 0b10010001000000000000000010000100000,
          'fMyB' : 0b10010001000000000000000001000010100,
          'fMyZ' : 0b10010001000000000000000001000100100,
          'fMZD' : 0b10010001000000100010000000100100000,
          'fMZS' : 0b10010001000000100010000000100010000,
          'fS' : 0b10000000000000000000000000010010000,
          'fSD' : 0b10000100100000000000000000001100000,
          'fSS' : 0b10000100100000000000000000001010000,
          'fSW' : 0b10000100100000000000000000010100001,
          'fyB' : 0b10000000000000000000000000100010000,
          'fyBD' : 0b10001000101000000000000000010100000,
          'fyBS' : 0b10001000101000000000000000010010000,
          'fyZ' : 0b10000000000000000000000000100100000,
          'fyZD' : 0b10001001001000000000000000010100000,
          'fyZS' : 0b10001001001000000000000000010010000,
          'fZH' : 0b10001001000010000000000000100010000,
          'mBW' : 0b01001000100000000000000000100100001,
          'mD' : 0b01000000000000000000000000010100000,
          'mDD' : 0b01000101000000000000000000001100000,
          'mDH' : 0b01000101000000000000000000010010001,
          'mDS' : 0b01000101000000000000000000001010000,
          'meB' : 0b01000000000000000000000000100011000,
          'meBD' : 0b01001000110000000000000000010100000,
          'meBS' : 0b01001000110000000000000000010010000,
          'meZ' : 0b01000000000000000000000000100101000,
          'meZD' : 0b01001001010000000000000000010100000,
          'meZS' : 0b01001001010000000000000000010010000,
          'mF' : 0b01000000000000000000000001000010000,
          'mFBD' : 0b01010000100000100001000000100100000,
          'mFBS' : 0b01010000100000100001000000100010000,
          'mFeB' : 0b01010000100000000000000001000011000,
          'mFeZ' : 0b01010000100000000000000001000101000,
          'mFF' : 0b01010000100000000000000010000010000,
          'mFM' : 0b01010000100000000000000010000100000,
          'mFyB' : 0b01010000100000000000000001000010100,
          'mFyZ' : 0b01010000100000000000000001000100100,
          'mFZD' : 0b01010000100000100010000000100100000,
          'mFZS' : 0b01010000100000100010000000100010000,
          'mM' : 0b01000000000000000000000001000100000,
          'mMBD' : 0b01010001000000100001000000100100000,
          'mMBS' : 0b01010001000000100001000000100010000,
          'mMeB' : 0b01010001000000000000000001000011000,
          'mMeZ' : 0b01010001000000000000000001000101000,
          'mMF' : 0b01010001000000000000000010000010000,
          'mMM' : 0b01010001000000000000000010000100000,
          'mMyB' : 0b01010001000000000000000001000010100,
          'mMyZ' : 0b01010001000000000000000001000100100,
          'mMZD' : 0b01010001000000100010000000100100000,
          'mMZS' : 0b01010001000000100010000000100010000,
          'mS' : 0b01000000000000000000000000010010000,
          'mSD' : 0b01000100100000000000000000001100000,
          'mSS' : 0b01000100100000000000000000001010000,
          'mSW' : 0b01000100100000000000000000010100001,
          'mW' : 0b01000000000000000000000000100100001,
          'mWB' : 0b01001001000010000000000000100100000,
          'mWF' : 0b01001001000010000000000001000010000,
          'mWM' : 0b01001001000010000000000001000100000,
          'mWZ' : 0b01001001000010000000000000100100001,
          'myB' : 0b01000000000000000000000000100010100,
          'myBD' : 0b01001000101000000000000000010100000,
          'myBS' : 0b01001000101000000000000000010010000,
          'myZ' : 0b01000000000000000000000000100100100,
          'myZD' : 0b01001001001000000000000000010100000,
          'myZS' : 0b01001001001000000000000000010010000,
          'mZH' : 0b01001001000000000000000000100010001,
         }
# creating the binary data model of the paradigm

def create_model(data):
    data.pop(0)
    model_dict = []
#    i = 0
    for x in data:
        dictline = {}
        for k1, v1 in x.items():
#            print(k1)
            a = "Language"
            b = "ISOCODE"
            c = "qqq"
            if (k1 == a):
                dictline[k1] = v1
                continue
            elif (k1 == b):
                dictline[k1] = v1
                continue
            elif (v1 == c):
              dictline[k1] = "?????????????????????????????????????"
              continue
            equivalence = canon.get(k1)
            #print(equivalence)
            for k2, v2 in x.items():
                #  print(k2)
                  if (k2 == a):
                    continue
                  elif (k2 == b):
                    continue
                  elif (k2 == k1):
                    continue
                  elif (v2 == v1):
                    equivalence = equivalence|canon.get(k2)
                  dictline[k1] = bin(equivalence)
                  print(dictline)
        model_dict.append(dictline)
    return(model_dict)
       
                    

def make_nexus(model_dict):
    sequence = []
    n = NexusWriter()
    dictline = {}
    for dictline in model_dict:
        lang = dictline.get('Language')
        isocode = dictline.get('ISOCODE')
        kvec = ""
        for k in sorted(dictline):
            #print(k)
            if (k == 'Language' or k == 'ISOCODE'):
                continue
            else:
                k = repr(dictline.get(k))
                krem = k.replace("'", "")
                kmod = krem[2:36]
                kvec = kvec + kmod.zfill(35)
# label is number of elements in vector, number of taxa and a random 1
        label = "%s_%s_%d" % ("756", "51", 1)
        n.add(lang, label, kvec)
    return n

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Generates a nexus')
    # positional
    parser.add_argument("xlsfile", help="Excel file")
    parser.add_argument("outfile", help="Output file")
    args = parser.parse_args()
    
    if os.path.isfile(args.outfile):
        raise IOError(
            "Output file %s already exists, please rename" % args.outfile
        )
    
    if not os.path.isfile(args.xlsfile):
        raise IOError(
            "Unable to find input xlsx file %s" % args.xlsfile
        )
    
    data = get_data(args.xlsfile)
    model_dict = create_model(data)
    nex = make_nexus(model_dict)
    nex.write_to_file(args.outfile, charblock=True)